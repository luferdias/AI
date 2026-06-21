#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Monta um Excel com os dados da memoria de massa IFES 2022/2023 e
GRAFICOS NATIVOS do Excel (LineChart/AreaChart) construidos a partir dos
dados das celulas - nao sao imagens."""
import datetime as dt
import openpyxl
from openpyxl.chart import LineChart, AreaChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties

SRC = "/root/.claude/uploads/018b0349-7f4a-55f8-b0f5-025b5a844733/619d2718-memoria_massa_ifes_2022_2023_planilha.xlsx"
OUTXLSX = "/home/user/AI/memoria_massa_ifes_2022_2023_com_graficos.xlsx"

COL = {"a": "1F77B4", "b": "D9622B", "c": "2A9D8F", "d": "8856A7"}
GRAY = "EDEDED"
ORANGE = "FBE2BD"
DARKRED = "8B0000"
SHEET = {"a": "Planilha a)", "b": "Planilha b)", "c": "Planilha c)", "d": "Planilha d)"}
LABELS = {
    "a": "a) Férias no verão",
    "b": "b) Atividades acadêmicas regulares no verão",
    "c": "c) Atividades acadêmicas regulares no inverno",
    "d": "d) Férias no inverno",
}
TITLES_IND = {
    "a": "a) Férias no verão - 09 a 15 de janeiro de 2023",
    "b": "b) Atividades acadêmicas regulares no verão - 13 a 19 de março de 2023",
    "c": "c) Atividades acadêmicas regulares no inverno - 07 a 13 de agosto de 2023",
    "d": "d) Férias no inverno - 17 a 23 de julho de 2023",
}

print("Lendo origem...")
wb_src = openpyxl.load_workbook(SRC, data_only=True)

# ----- carregar (data, demanda) de cada periodo
def load(period):
    ws = wb_src[SHEET[period]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        out.append((r[0], None if r[4] is None else float(r[4])))
    return out
DATA = {p: load(p) for p in "abcd"}

# ----- copiar todas as abas de dados para o novo arquivo
print("Copiando dados...")
wb = openpyxl.Workbook()
wb.remove(wb.active)
for name in wb_src.sheetnames:
    if name == "Gráficos":
        continue
    ws_in = wb_src[name]
    ws_out = wb.create_sheet(title=name)
    for row in ws_in.iter_rows(values_only=True):
        ws_out.append(list(row))

# ----- colunas auxiliares (bandas + pico) em cada Planilha
# A=Data(1) ... E=Demanda(5). Helpers: P=16 FimSemana, Q=17 Ponta, R=18 Pico
def add_helpers(period):
    ws = wb[SHEET[period]]
    rows = DATA[period]
    pk_i = max(range(len(rows)), key=lambda i: (rows[i][1] if rows[i][1] is not None else -1))
    ws.cell(row=1, column=16, value="Fim de semana")
    ws.cell(row=1, column=17, value="Horário de ponta")
    ws.cell(row=1, column=18, value="Pico")
    for i, (d, v) in enumerate(rows):
        r = i + 2
        wd = d.weekday()
        ws.cell(row=r, column=16, value=600 if wd >= 5 else 0)
        ws.cell(row=r, column=17, value=600 if (wd < 5 and 18 <= d.hour < 21) else 0)
        ws.cell(row=r, column=18, value=(v if i == pk_i else None))
    return len(rows), pk_i, rows[pk_i]

HELP = {p: add_helpers(p) for p in "abcd"}

# ----- abas auxiliares de calculo: Perfil medio e Comparacao
# Perfil medio dos dias uteis: media por slot de 15min (seg-sex)
ws_perf = wb.create_sheet("Aux_Perfil")
ws_perf.append(["Hora", LABELS["a"], LABELS["b"], LABELS["c"], LABELS["d"], "Ponta"])
prof = {p: [0.0] * 96 for p in "abcd"}
cnt = {p: [0] * 96 for p in "abcd"}
for p in "abcd":
    for d, v in DATA[p]:
        if d.weekday() < 5 and v is not None:
            s = d.hour * 4 + d.minute // 15
            prof[p][s] += v
            cnt[p][s] += 1
for s in range(96):
    hora = s * 0.25
    vals = [(prof[p][s] / cnt[p][s]) if cnt[p][s] else None for p in "abcd"]
    ponta = 600 if 18 <= hora < 21 else 0
    ws_perf.append([round(hora, 2)] + [None if v is None else round(v, 2) for v in vals] + [ponta])

# Comparacao das 4 curvas (semana inteira, indices alinhados)
ws_cmp = wb.create_sheet("Aux_Comparacao")
ws_cmp.append(["Intervalo", LABELS["a"], LABELS["b"], LABELS["c"], LABELS["d"], "Fim de semana"])
n = min(len(DATA[p]) for p in "abcd")
dias_lbl = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
for i in range(n):
    # rotulo do dia apenas no inicio de cada dia, senao vazio
    lbl = dias_lbl[i // 96] if i % 96 == 0 else ""
    fimsem = 600 if (i // 96) >= 5 else 0
    vals = [DATA[p][i][1] for p in "abcd"]
    ws_cmp.append([lbl] + vals + [fimsem])

# ================================================================ helpers de estilo
def style_line(series, hexcolor, width=19050):
    lp = LineProperties(solidFill=hexcolor, w=width)
    series.graphicalProperties = GraphicalProperties()
    series.graphicalProperties.line = lp
    series.smooth = False
    series.marker = Marker(symbol="none")

def style_band(series, hexcolor):
    series.graphicalProperties = GraphicalProperties()
    series.graphicalProperties.solidFill = hexcolor
    series.graphicalProperties.line = LineProperties(noFill=True)

def style_peak(series):
    series.graphicalProperties = GraphicalProperties()
    series.graphicalProperties.line = LineProperties(noFill=True)
    m = Marker(symbol="circle", size=8)
    m.graphicalProperties = GraphicalProperties()
    m.graphicalProperties.solidFill = DARKRED
    m.graphicalProperties.line = LineProperties(solidFill=DARKRED)
    series.marker = m

def base_axes(chart, ytitle, xtitle, title):
    chart.title = title
    chart.y_axis.title = ytitle
    chart.x_axis.title = xtitle
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 600
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.height = 9
    chart.width = 30

# ================================================================ aba de graficos
ws_g = wb.create_sheet("Gráficos")
ws_g["A1"] = "Gráficos nativos do Excel construídos a partir dos dados (demanda em kW)"

def chart_individual(period, anchor):
    ws_data = wb[SHEET[period]]
    nrows, pk_i, (pk_d, pk_v) = HELP[period]
    last = nrows + 1
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=last)

    # area com as bandas (fundo)
    area = AreaChart()
    area.grouping = "standard"
    for col, color in ((16, GRAY), (17, ORANGE)):
        ref = Reference(ws_data, min_col=col, min_row=1, max_row=last)
        s = Series(ref, title_from_data=True)
        style_band(s, color)
        area.series.append(s)
    area.set_categories(cats)

    # linha de demanda
    line = LineChart()
    ref = Reference(ws_data, min_col=5, min_row=1, max_row=last)
    s = Series(ref, title_from_data=True)
    style_line(s, COL[period])
    line.series.append(s)
    # marcador de pico
    refp = Reference(ws_data, min_col=18, min_row=1, max_row=last)
    sp = Series(refp, title_from_data=True)
    style_peak(sp)
    line.series.append(sp)
    line.set_categories(cats)

    area += line
    base_axes(area, "Demanda (kW)", "Data e hora - intervalos de 15 min",
              f"{TITLES_IND[period]}  |  Pico {pk_v:.2f} kW em {pk_d.strftime('%d/%m %H:%M')}")
    area.legend = None
    ws_g.add_chart(area, anchor)

def chart_perfil(anchor):
    last = 97
    cats = Reference(ws_perf, min_col=1, min_row=2, max_row=last)
    area = AreaChart()
    area.grouping = "standard"
    ref = Reference(ws_perf, min_col=6, min_row=1, max_row=last)
    s = Series(ref, title_from_data=True)
    style_band(s, ORANGE)
    area.series.append(s)
    area.set_categories(cats)

    line = LineChart()
    for k, p in enumerate("abcd"):
        ref = Reference(ws_perf, min_col=2 + k, min_row=1, max_row=last)
        s = Series(ref, title_from_data=True)
        style_line(s, COL[p], width=22000)
        line.series.append(s)
    line.set_categories(cats)

    area += line
    base_axes(area, "Demanda média (kW)", "Hora do dia",
              "Perfil médio dos dias úteis - comparação sazonal e acadêmica")
    area.legend.position = "r"
    ws_g.add_chart(area, anchor)

def chart_comparacao(anchor):
    last = n + 1
    cats = Reference(ws_cmp, min_col=1, min_row=2, max_row=last)
    area = AreaChart()
    area.grouping = "standard"
    ref = Reference(ws_cmp, min_col=6, min_row=1, max_row=last)
    s = Series(ref, title_from_data=True)
    style_band(s, GRAY)
    area.series.append(s)
    area.set_categories(cats)

    line = LineChart()
    for k, p in enumerate("abcd"):
        ref = Reference(ws_cmp, min_col=2 + k, min_row=1, max_row=last)
        s = Series(ref, title_from_data=True)
        style_line(s, COL[p], width=12700)
        line.series.append(s)
    line.set_categories(cats)

    area += line
    base_axes(area, "Demanda (kW)", "Tempo decorrido na semana",
              "Comparação das quatro curvas - mesma escala de demanda")
    area.legend.position = "r"
    ws_g.add_chart(area, anchor)

print("Construindo graficos nativos...")
chart_individual("a", "A3")
chart_perfil("A22")
chart_comparacao("A41")
chart_individual("d", "A60")
chart_individual("c", "A79")
chart_individual("b", "A98")

print("Salvando", OUTXLSX)
wb.save(OUTXLSX)
print("OK")
