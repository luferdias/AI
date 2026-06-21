#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Reproduz as figuras de demanda da memoria de massa IFES 2022/2023
e monta um novo arquivo Excel com os dados + as figuras geradas."""
import datetime as dt
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import FancyBboxPatch
import openpyxl
from openpyxl.drawing.image import Image as XLImage

SRC = "/root/.claude/uploads/018b0349-7f4a-55f8-b0f5-025b5a844733/619d2718-memoria_massa_ifes_2022_2023_planilha.xlsx"
OUTDIR = "/home/user/AI/figuras"
OUTXLSX = "/home/user/AI/memoria_massa_ifes_2022_2023_com_graficos.xlsx"

import os
os.makedirs(OUTDIR, exist_ok=True)

# ---------------------------------------------------------------- cores
COL = {"a": "#1f77b4", "b": "#d9622b", "c": "#2a9d8f", "d": "#8856a7"}
ORANGE = "#f5c56e"     # faixa horario de ponta
GRAY = "#e6e6e6"       # faixa fim de semana
DARKRED = "#8b0000"

LABELS = {
    "a": "a) Férias no verão",
    "b": "b) Atividades acadêmicas regulares no verão",
    "c": "c) Atividades acadêmicas regulares no inverno",
    "d": "d) Férias no inverno",
}
TITLES_IND = {
    "a": "a) Férias no verão - 09 a 15 de janeiro de 2023",
    "c": "c) Atividades acadêmicas regulares no inverno - 07 a 13 de agosto de 2023",
    "d": "d) Férias no inverno - 17 a 23 de julho de 2023",
    "b": "b) Atividades acadêmicas regulares no verão - 13 a 19 de março de 2023",
}

# ---------------------------------------------------------------- leitura
print("Lendo planilha de origem...")
wb_src = openpyxl.load_workbook(SRC, data_only=True)
SHEET = {"a": "Planilha a)", "b": "Planilha b)", "c": "Planilha c)", "d": "Planilha d)"}

def load(period):
    ws = wb_src[SHEET[period]]
    datas, dem = [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        datas.append(r[0])
        dem.append(r[4])
    datas = np.array(datas, dtype="datetime64[m]")
    dem = np.array([np.nan if v is None else float(v) for v in dem])
    return datas, dem

DATA = {p: load(p) for p in "abcd"}

# ---------------------------------------------------------------- estilo geral
plt.rcParams.update({
    "font.size": 11,
    "axes.titlesize": 14,
    "axes.labelsize": 12,
    "axes.grid": True,
    "grid.color": "#d9d9d9",
    "grid.linewidth": 0.8,
    "axes.edgecolor": "#333333",
    "figure.facecolor": "white",
})

def fmt_pico(v):
    return f"{v:.2f}"

# ================================================================ figuras individuais (a, c, d)
def fig_individual(period):
    datas, dem = DATA[period]
    t = datas.astype("datetime64[s]").astype("O")  # datetime objects
    start = dt.datetime(t[0].year, t[0].month, t[0].day)
    end = start + dt.timedelta(days=7)

    fig, ax = plt.subplots(figsize=(14, 4.8))
    ax.plot(t, dem, color=COL[period], lw=1.2)

    ax.set_ylim(0, 600)
    ax.set_xlim(start, end)
    ax.set_title(TITLES_IND[period])
    ax.set_xlabel("Data e hora - intervalos de 15 min")
    ax.set_ylabel("Demanda (kW)")

    # faixa fim de semana (sab + dom)
    sat = start + dt.timedelta(days=5)
    ax.axvspan(sat, end, color=GRAY, alpha=0.9, zorder=0)

    # faixas horario de ponta 18-21h nos dias uteis (seg-sex)
    for d in range(5):
        day = start + dt.timedelta(days=d)
        ax.axvspan(day + dt.timedelta(hours=18), day + dt.timedelta(hours=21),
                   color=ORANGE, alpha=0.55, zorder=0)

    # eixo x: um tick por dia, rotulo "Mon\n09/01"
    ax.xaxis.set_major_locator(mdates.DayLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%a\n%d/%m"))
    ax.tick_params(axis="x", length=0)

    # ponto de pico
    i = int(np.nanargmax(dem))
    pv = dem[i]
    pt = t[i]
    ax.plot([pt], [pv], "o", color=DARKRED, ms=9, zorder=5)
    txt = f"Pico {fmt_pico(pv)} kW\n{pt.strftime('%d/%m %H:%M')}"
    ax.annotate(
        txt, xy=(pt, pv),
        xytext=(pt + dt.timedelta(hours=8), pv - 80),
        color=DARKRED, fontsize=10,
        bbox=dict(boxstyle="round,pad=0.4", fc="white", ec=DARKRED, lw=1.5),
        arrowprops=dict(arrowstyle="-", color=DARKRED, lw=1.5),
    )

    # textos auxiliares
    ax.text(0.005, 0.03, "Faixa cinza: fim de semana | Faixa laranja: horário de ponta (18h-21h)",
            transform=ax.transAxes, fontsize=9, color="#7a7a7a", ha="left", va="bottom")
    ax.text(0.995, 0.95, "Escala comum: 0 a 600 kW",
            transform=ax.transAxes, fontsize=10, color="#7a7a7a", ha="right", va="top")

    ax.grid(True, axis="both", zorder=1)
    fig.tight_layout()
    path = f"{OUTDIR}/fig_{period}.png"
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path

# ================================================================ perfil medio dias uteis
def fig_perfil_medio():
    fig, ax = plt.subplots(figsize=(14, 4.8))
    x = np.arange(96) * 0.25  # horas do dia
    for p in "abcd":
        datas, dem = DATA[p]
        tt = datas.astype("datetime64[s]").astype("O")
        # media por slot de 15 min apenas nos dias uteis (seg-sex)
        acc = np.zeros(96)
        cnt = np.zeros(96)
        for k, d in enumerate(tt):
            if d.weekday() < 5 and not np.isnan(dem[k]):
                slot = d.hour * 4 + d.minute // 15
                acc[slot] += dem[k]
                cnt[slot] += 1
        prof = np.where(cnt > 0, acc / np.maximum(cnt, 1), np.nan)
        ax.plot(x, prof, color=COL[p], lw=1.6, label=LABELS[p])

    ax.set_ylim(0, 600)
    ax.set_xlim(0, 24)
    ax.set_xticks(range(0, 25, 3))
    ax.set_title("Perfil médio dos dias úteis - comparação sazonal e acadêmica")
    ax.set_xlabel("Hora do dia")
    ax.set_ylabel("Demanda média (kW)")
    ax.axvspan(18, 21, color=ORANGE, alpha=0.45, zorder=0)
    ax.legend(loc="upper right", framealpha=0.95)
    fig.tight_layout()
    path = f"{OUTDIR}/fig_perfil_medio.png"
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path

# ================================================================ comparacao 4 curvas (semana)
def fig_comparacao():
    fig, ax = plt.subplots(figsize=(14, 4.8))
    n = 672
    x = np.arange(n)
    for p in "abcd":
        _, dem = DATA[p]
        y = dem[:n]
        ax.plot(x[:len(y)], y, color=COL[p], lw=1.0, label=LABELS[p])

    ax.set_ylim(0, 600)
    ax.set_xlim(0, n - 1)
    ax.set_title("Comparação das quatro curvas - mesma escala de demanda")
    ax.set_xlabel("Tempo decorrido na semana")
    ax.set_ylabel("Demanda (kW)")

    dias = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    ax.set_xticks([d * 96 + 48 for d in range(7)])
    ax.set_xticklabels(dias)
    ax.tick_params(axis="x", length=0)

    # faixa fim de semana (sab+dom => indices 5*96 ate o fim)
    ax.axvspan(5 * 96, n - 1, color=GRAY, alpha=0.9, zorder=0)

    ax.text(0.995, 0.03, "Todos os períodos: eixo Y fixado entre 0 e 600 kW",
            transform=ax.transAxes, fontsize=9, color="#7a7a7a", ha="right", va="bottom")
    ax.legend(loc="upper right", framealpha=0.95)
    fig.tight_layout()
    path = f"{OUTDIR}/fig_comparacao.png"
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path

print("Gerando figuras...")
paths = {}
paths["a"] = fig_individual("a")
paths["c"] = fig_individual("c")
paths["d"] = fig_individual("d")
paths["b"] = fig_individual("b")
paths["perfil"] = fig_perfil_medio()
paths["comp"] = fig_comparacao()
print("Figuras geradas:", paths)

# ================================================================ montar novo xlsx com dados + figuras
print("Copiando dados para o novo arquivo...")
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

for name in wb_src.sheetnames:
    ws_in = wb_src[name]
    ws_out = wb_out.create_sheet(title=name)
    if name == "Gráficos":
        continue  # vamos preencher com imagens
    for row in ws_in.iter_rows(values_only=True):
        ws_out.append(list(row))

# Inserir figuras na aba Gráficos
ws_g = wb_out["Gráficos"]
ws_g["A1"] = "Figuras reproduzidas a partir dos dados de memória de massa (demanda em kW)"
order = [("a", "Figura a)"), ("perfil", "Perfil médio dos dias úteis"),
         ("comp", "Comparação das quatro curvas"), ("d", "Figura d)"),
         ("c", "Figura c)"), ("b", "Figura b)")]
row = 3
for key, _ in order:
    img = XLImage(paths[key])
    # escala para caber bem
    img.width = int(img.width * 0.9)
    img.height = int(img.height * 0.9)
    ws_g.add_image(img, f"A{row}")
    row += 26  # espaco vertical entre imagens

print("Salvando", OUTXLSX)
wb_out.save(OUTXLSX)
print("OK")
EOF